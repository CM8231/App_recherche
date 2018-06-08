# -*- coding: utf-8 -*-

import os
import random
import time
import re
from glob import glob
from pygame import mixer
import winsound

from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import QThread, pyqtSignal, QObject
# from selenium.webdriver.support.ui import WebDriverWait
# from selenium.webdriver.support.expected_conditions import staleness_of
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font, colors
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import WebDriverException
from selenium.common import exceptions
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait



class UiMainWindow(QtWidgets.QMainWindow):
    def __init__(self, dim):
        super().__init__()
        ####### Définition des variables utiles  #############
        self.dim = dim  # Récupère les dimensions de l'écran
        self.setObjectName("main_window")
        width = 452
        self.setGeometry(0, 38, width, self.dim.height() - 90)
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap("Adentis_icon.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        # Icon dans le dossier de l'application
        self.setWindowIcon(icon)
        self.setStyleSheet(
            "background-color: qlineargradient(spread:reflect, x1:0.513, y1:0, x2:0.517, y2:0.511, "
            "stop:0 rgba(0, 158, 255, 255), stop:0.283582 rgba(255, 255, 255, 255));")

        self.central_widget = QtWidgets.QWidget(self)
        self.central_widget.setObjectName("central_widget")
        self.setCentralWidget(self.central_widget)

        ############# Définition de la grille du layout ##############
        self.gridLayout = QtWidgets.QGridLayout(self.central_widget)
        self.gridLayout.setContentsMargins(10, 35, 10, 10)
        self.gridLayout.setObjectName("gridLayout")

        ############## statusbar ################
        self.status_bar = QtWidgets.QStatusBar(self.central_widget)
        self.status_bar.setObjectName("status_bar")
        self.setStatusBar(self.status_bar)

        ############# menu bar #############
        menu = QtWidgets.QMenuBar(self.central_widget)
        menu.setStyleSheet("background-color: qlineargradient(spread:reflect, x1:0.518, y1:0, x2:0.50705, "
                           "y2:0.517, stop:0 rgba(207, 211, 214, 255), stop:0.995025 rgba(250, 250, 250, 255));\n "
                           "selection-color: rgb(0, 0, 0); \n selection-background-color: rgb(0, 170, 255);")
        optionMenu = menu.addMenu('&Options')
        self.audioAct = QtWidgets.QAction('Audio', self.central_widget)
        self.audioAct.setCheckable(True)
        self.audioAct.setStatusTip('Emettre un son à la fin de la recherche')
        self.audioAct.setChecked(False)
        # self.audioAct.triggered.connect(self.toggleMenu)
        # exitAct = QtWidgets.QAction('Quitter',self.central_widget)
        # exitAct.setShortcut('Ctrl+Q')
        # exitAct.setStatusTip("Quitte l'application")
        # exitAct.triggered.connect(app.quit)
        # optionMenu.addAction(exitAct)
        self.progress_auto_quit_Act = QtWidgets.QAction('Auto-fermeture Progression', self.central_widget)
        self.progress_auto_quit_Act.setCheckable(True)
        self.progress_auto_quit_Act.setStatusTip('Ferme la barre de progression à la fin de la recherche')
        self.progress_auto_quit_Act.setChecked(False)
        # self.progress_auto_quit_Act.triggered.connect(self.toggleMenu)
        optionMenu.addAction(self.audioAct)
        optionMenu.addAction(self.progress_auto_quit_Act)

        # Ajoute un menu secondaire pour les options de recherche
        recherche_option_menu = QtWidgets.QMenu('Options de Recherche', self.central_widget)
        recherche_option_menu.setStyleSheet(
            "background-color: qlineargradient(spread:reflect, x1:0.518, y1:0, x2:0.50705, "
            "y2:0.517, stop:0 rgba(207, 211, 214, 255), stop:0.995025 rgba(250, 250, 250, 255));\n "
            "selection-color: rgb(0, 0, 0); \n selection-background-color: rgb(0, 170, 255);")

        # Ajoute un menu tertiaire pour les types de poste recherché
        recherche_option_type_poste_menu = QtWidgets.QMenu('Type de Poste', self.central_widget)
        recherche_option_type_poste_menu.setStyleSheet(
            "background-color: qlineargradient(spread:reflect, x1:0.518, y1:0, x2:0.50705, "
            "y2:0.517, stop:0 rgba(207, 211, 214, 255), stop:0.995025 rgba(250, 250, 250, 255));\n "
            "selection-color: rgb(0, 0, 0); \n selection-background-color: rgb(0, 170, 255);")
        # Crée les action déterminant le type de poste
        self.recherche_type_both_Act = QtWidgets.QAction('Actuel + Auparavant', self.central_widget)
        self.recherche_type_both_Act.setCheckable(True)
        self.recherche_type_both_Act.setStatusTip("Recherche globale sur l'entreprise: Poste Actuel et Passé")
        self.recherche_type_both_Act.setChecked(True)
        self.recherche_type_both_Act.setObjectName("Actuel+Auparavant")
        self.recherche_type_both_Act.triggered[bool].connect(self.select_recherche_type)

        self.recherche_type_actuel_Act = QtWidgets.QAction('Actuel', self.central_widget)
        self.recherche_type_actuel_Act.setCheckable(True)
        self.recherche_type_actuel_Act.setStatusTip("Recherche contrainte sur l'entreprise: Poste Actuel uniquement")
        self.recherche_type_actuel_Act.setChecked(False)
        self.recherche_type_actuel_Act.setObjectName("Actuel")
        self.recherche_type_actuel_Act.triggered[bool].connect(self.select_recherche_type)

        self.recherche_type_auparavant_Act = QtWidgets.QAction('Auparavant', self.central_widget)
        self.recherche_type_auparavant_Act.setCheckable(True)
        self.recherche_type_auparavant_Act.setStatusTip("Recherche contrainte sur l'entreprise: Poste Passé uniquement")
        self.recherche_type_auparavant_Act.setChecked(False)
        self.recherche_type_auparavant_Act.setObjectName("Auparavant")
        self.recherche_type_auparavant_Act.triggered[bool].connect(self.select_recherche_type)

        # Détermine si on récupère les coordonnées dans les profils
        self.optionProfil = QtWidgets.QAction('Coordonnées', self.central_widget)
        self.optionProfil.setCheckable(True)
        self.optionProfil.setStatusTip("Récupérer les coordonnées disponibles sur le profil")
        self.optionProfil.setChecked(True)

        # Construit l'architecture des menus
        recherche_option_menu.addAction(self.optionProfil)
        recherche_option_type_poste_menu.addAction(self.recherche_type_both_Act)
        recherche_option_type_poste_menu.addAction(self.recherche_type_actuel_Act)
        recherche_option_type_poste_menu.addAction(self.recherche_type_auparavant_Act)

        recherche_option_menu.addMenu(recherche_option_type_poste_menu)
        optionMenu.addMenu(recherche_option_menu)

        self.recherche_type = 'Actuel+Auparavant'  # Valeur par défaut

        ############## Choix du navigateur ######################
        self.titre_navigateur = QtWidgets.QLabel(self.central_widget)
        self.titre_navigateur.setStyleSheet("font: 87 11pt \"Arial Black\";\n"
                                            "background-color: qlineargradient(spread:pad, x1:0.518, y1:0, "
                                            "x2:0.517, y2:1, stop:0 rgba(186, 230, 255, 0), "
                                            "stop:0.995025 rgba(250, 250, 250, 0));")
        # StyleSheet background-color  pour transparence
        self.titre_navigateur.setObjectName("titre_navigateur")
        self.gridLayout.addWidget(self.titre_navigateur, 0, 2, 1, 1)

        self.combo_box = QtWidgets.QComboBox(self.central_widget)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(11)
        font.setBold(True)
        font.setWeight(75)
        self.combo_box.setFont(font)
        self.combo_box.setStyleSheet(
            "background-color: qlineargradient(spread:reflect, x1:0.517413, y1:0, x2:0.512, y2:0.494, "
            "stop:0 rgba(192, 192, 192, 255), stop:1 rgba(255, 255, 255, 255));\n"
            "border-color: rgb(15, 15, 15);\n"
            "selection-background-color: qlineargradient(spread:reflect, x1:0.493, y1:0.528, x2:0.497, y2:0, "
            "stop:0 rgba(192, 192, 192, 255), stop:1 rgba(255, 255, 255, 255));\n"
            "selection-color: rgb(0,0,0);")
        self.combo_box.setIconSize(QtCore.QSize(30, 30))
        self.combo_box.setObjectName("combo_box")
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap("firefox1600.png"), QtGui.QIcon.Normal, QtGui.QIcon.On)
        # Icon dans le dossier de l'application
        self.combo_box.addItem(icon, "")
        icon1 = QtGui.QIcon()
        icon1.addPixmap(QtGui.QPixmap("chrome1600.png"), QtGui.QIcon.Normal, QtGui.QIcon.On)
        # Icon dans le dossier de l'application
        self.combo_box.addItem(icon1, "")

        self.navigateur = "Firefox"  # Fixe le navigateur sur Firefox par défaut
        self.combo_box.setCurrentIndex(0)
        self.combo_box.activated[str].connect(self.nav_activated)
        # Quand combo_box activé récupère l'onglet choisit par l'intermédiaire de NavActivated
        self.combo_box.setToolTip('Choisissez le navigateur installé sur votre ordinateur')
        self.gridLayout.addWidget(self.combo_box, 1, 2, 1, 1)

        ################## Identification Linkedin ################
        self.titre_id_linkedin = QtWidgets.QLabel(self.central_widget)
        font = QtGui.QFont()
        font.setFamily("Arial Black")
        font.setPointSize(11)
        font.setBold(False)
        font.setItalic(False)
        font.setWeight(10)
        self.titre_id_linkedin.setFont(font)
        self.titre_id_linkedin.setStyleSheet("font: 87 11pt \"Arial Black\";\n"
                                             "background-color: qlineargradient(spread:pad, x1:0.518, y1:0, x2:0.517, "
                                             "y2:1, stop:0 rgba(186, 230, 255, 0), stop:0.995025 rgba(250, 250, 250, 0)"
                                             ");")
        self.titre_id_linkedin.setObjectName("titre_id_linkedin")
        self.gridLayout.addWidget(self.titre_id_linkedin, 0, 1, 1, 1)

        ## Identifiant linkedin ##
        self.identifiant = QtWidgets.QLineEdit(self.central_widget)

        self.identifiant.setStyleSheet(
            "background-color: qlineargradient(spread:pad, x1:0.493, y1:0.528, x2:0.497, y2:0, "
            "stop:0.134328 rgba(209, 209, 209, 255), stop:1 rgba(255, 255, 255, 255));\n"
            "font: 12pt \"Arial\";\n"
            "alternate-background-color: rgb(255, 255, 255);")

        self.identifiant.setToolTip(
            "<html><head/><body><p align=\"justify\"><span style=\" "
            "font-size: 9pt;\">Entrez votre adresse mail de connexion à Linkedin</span></p></body></html>")
        self.identifiant.setObjectName("identifiant")
        self.identifiant.setPlaceholderText('Adresse Email')
        # Affiche un texte quand pas encore d'input dans le LineEdit
        self.gridLayout.addWidget(self.identifiant, 1, 1, 1, 1)

        ## Mot de passe linkedin ##
        self.password = QtWidgets.QLineEdit(self.central_widget)
        self.password.setStyleSheet(
            "background-color: qlineargradient(spread:pad, x1:0.493, y1:0.528, x2:0.497, y2:0, "
            "stop:0.134328 rgba(209, 209, 209, 255), stop:1 rgba(255, 255, 255, 255));\n"
            "font: 12pt \"Arial\";")
        self.password.setToolTip(
            "<html><head/><body><p align=\"justify\"><span style=\" "
            "font-size: 9pt;\">Entrez votre mot de passe de connexion à Linkedin</span></p></body></html>")
        self.password.setObjectName("password")
        self.password.setEchoMode(2)  # EchoMode= 2 Pour afficher l'input en *****
        self.password.setPlaceholderText('Mot de passe')  # Affiche un texte quand pas encore d'input dans le LineEdit
        self.gridLayout.addWidget(self.password, 2, 1, 1, 1)

        ################  Recherche #####################
        ##Grille secondaire pour paramètre de recherche ##
        self.gridLayout_2 = QtWidgets.QGridLayout()
        self.gridLayout_2.setObjectName("gridLayout_2")

        self.titre_recherche = QtWidgets.QLabel(self.central_widget)
        self.titre_recherche.setStyleSheet("font: 87 10pt \"Arial Black\"\n;"
                                           "background-color: qlineargradient(spread:pad, x1:0.518, y1:0, x2:0.517, "
                                           "y2:1, stop:0 rgba(186, 230, 255, 0), "
                                           "stop:0.995025 rgba(250, 250, 250, 0));")
        self.titre_recherche.setObjectName("titre_recherche")
        self.gridLayout.addWidget(self.titre_recherche, 3, 1, 1, 1)

        ## Recherche nom entreprise ##
        # Titre
        self.society = QtWidgets.QLabel(self.central_widget)
        self.society.setStyleSheet("font: 9pt \"Arial\";\n"
                                   "text-decoration: underline;\n"
                                   "background-color: qlineargradient(spread:pad, x1:0.518, y1:0, x2:0.517, y2:1, "
                                   "stop:0 rgba(186, 230, 255, 0), stop:0.995025 rgba(250, 250, 250, 0));")
        self.society.setObjectName("society")
        self.gridLayout_2.addWidget(self.society, 1, 0, 1, 2)
        # Champ
        self.societe_champ = QtWidgets.QLineEdit(self.central_widget)
        self.societe_champ.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.societe_champ.setObjectName("societe_champ")
        self.societe_champ.setToolTip(
            "<html><head/><body><p align=\"justify\"><span style=\" "
            "font-size: 9pt;\">Nom de l'entreprise recherchée</span></p></body></html>")
        self.gridLayout_2.addWidget(self.societe_champ, 2, 0, 1, 2)

        ## Recherche localisation ##
        # Titre
        self.localisation = QtWidgets.QLabel(self.central_widget)
        self.localisation.setStyleSheet("font: 9pt \"Arial\";\n"
                                        "text-decoration: underline;\n"
                                        "background-color: qlineargradient(spread:pad, x1:0.518, y1:0, x2:0.517, y2:1, "
                                        "stop:0 rgba(186, 230, 255, 0), stop:0.995025 rgba(250, 250, 250, 0));")
        self.localisation.setObjectName("localisation")
        self.gridLayout_2.addWidget(self.localisation, 3, 0, 1, 2)
        # Champ
        self.localisation_champ = QtWidgets.QLineEdit(self.central_widget)
        self.localisation_champ.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.localisation_champ.setObjectName("localisation_champ")
        self.localisation_champ.setToolTip(
            "<html><head/><body><p align=\"justify\"><span style=\" "
            "font-size: 9pt;\">Localisation de l'entreprise ou mot clé additionnel</span></p></body></html>")
        self.gridLayout_2.addWidget(self.localisation_champ, 4, 0, 1, 2)

        ## liste des keywords pour les postes recherchés ##
        # Titre
        self.poste_keyword = QtWidgets.QLabel(self.central_widget)
        self.poste_keyword.setStyleSheet("font: 9pt \"Arial\";\n"
                                         "background-color: rgb(255, 255, 255);\n"
                                         "text-decoration: underline;")
        self.poste_keyword.setObjectName("poste_keyword")
        self.gridLayout_2.addWidget(self.poste_keyword, 5, 0, 1, 1)

        # Récupération des keywords et affichage dans un ListWidget
        self.listWidget = QListWidget(self.central_widget)
        self.listWidget.setObjectName("listWidget")
        self.listWidget.setToolTip(
            "<html><head/><body><p align=\"justify\"><span style=\" "
            "font-size: 9pt;\">Selectionnez les postes clés recherchés</span></p></body></html>")

        # Récupère liste de keywords dans fichier texte correspondant dans le dossier de l'application
        keywords = self.get_keywords('keywords-poste.txt')
        if keywords:
            for keyword in keywords:
                item = QtWidgets.QListWidgetItem()
                item.setFlags(
                    QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsUserCheckable | QtCore.Qt.ItemIsEnabled)
                item.setCheckState(QtCore.Qt.Checked)  # Coche par défaut tous les keywords extraits du fichier txt
                item.setText(keyword)
                self.listWidget.addItem(item)  # Ajoute les keywords dans le listWidget
        self.gridLayout_2.addWidget(self.listWidget, 6, 0, 1, 1)

        ## liste des keywords pour les mots clés à éviter ##
        # Titre
        self.poste_antikeyword = QtWidgets.QLabel(self.central_widget)
        self.poste_antikeyword.setStyleSheet("font: 9pt \"Arial\";\n"
                                             "background-color: rgb(255, 255, 255);\n"
                                             "text-decoration: underline;")
        self.poste_antikeyword.setObjectName("poste_antikeyword")
        self.gridLayout_2.addWidget(self.poste_antikeyword, 5, 1, 1, 1)

        # Récupération des keywords et affichage dans un ListWidget
        self.listWidget2 = QListWidget(self.central_widget)
        self.listWidget2.setObjectName("listWidget2")
        self.listWidget2.setToolTip(
            "<html><head/><body><p align=\"justify\"><span style=\" "
            "font-size: 9pt;\">Selectionnez les domaines à éviter</span></p></body></html>")

        # Récupère liste de keywords à éviter dans fichier texte correspondant dans le dossier de l'application
        keywords = self.get_keywords('antikeywords-poste.txt')
        if keywords:
            for keyword in keywords:
                item = QtWidgets.QListWidgetItem()
                item.setFlags(
                    QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsUserCheckable | QtCore.Qt.ItemIsEnabled)
                item.setCheckState(QtCore.Qt.Checked)  # Coche par défaut tous les keywords extraits du fichier txt
                item.setText(keyword)
                self.listWidget2.addItem(item)  # Ajoute les keywords dans le listWidget
        self.gridLayout_2.addWidget(self.listWidget2, 6, 1, 1, 1)

        # Ajoute la grille secondaire à la grille principale
        self.gridLayout.addLayout(self.gridLayout_2, 4, 1, 1, 2)

        ################# Chemin output ##############
        self.output_file = QtWidgets.QLineEdit(self.central_widget)
        self.output_file.setText('')
        self.output_file.setPlaceholderText('Chemin du fichier Excel')  # Texte quand pas d'input
        self.gridLayout.addWidget(self.output_file, 7, 1, 1, 1)

        self.parcourir = QtWidgets.QPushButton(self.central_widget)
        self.parcourir.setText('Parcourir')
        self.parcourir.resize(self.parcourir.sizeHint())
        self.gridLayout.addWidget(self.parcourir, 7, 2, 1, 1)
        self.parcourir.clicked.connect(self.output)  # Récupère le fichier excel d'output

        ############# Lancement du bot ###############
        self.lancement = QtWidgets.QPushButton(self.central_widget)
        self.lancement.setStyleSheet(
            "background-color: qlineargradient(spread:reflect, x1:0.493, y1:0.528, x2:0.497, y2:0, "
            "stop:0.104478 rgba(255, 255, 255, 255), stop:1 rgba(27, 192, 2, 255));\n"
            "selection-background-color: qlineargradient(spread:reflect, x1:0.493, y1:0.528, x2:0.497, y2:0, "
            "stop:0.134328 rgba(27, 192, 2, 255), stop:1 rgba(255, 255, 255, 255));\n"
            "font: 75 12pt \"MS Shell Dlg 2\";")
        self.lancement.setObjectName("lancement")
        self.lancement.clicked.connect(self.bot)  # Lance l'application de recherche linkedin
        self.gridLayout.addWidget(self.lancement, 9, 1, 1, 1)

        ################## Stoppe le bot ####################
        self.stop = QtWidgets.QPushButton(self.central_widget)
        self.stop.setToolTip("Stoppe la recherche")
        self.stop.setStyleSheet(
            "selection-background-color: qlineargradient(spread:reflect, x1:0.493, y1:0.528, x2:0.497, y2:0, "
            "stop:0 rgba(192, 0, 0, 255), stop:1 rgba(255, 255, 255, 255));\n"
            "background-color: qlineargradient(spread:reflect, x1:0.493, y1:0.528, x2:0.497, y2:0, "
            "stop:0 rgba(255, 255, 255, 255), stop:1 rgba(192, 0, 0, 255));\n"
            "font: 75 12pt \"MS Shell Dlg 2\";")
        self.stop.setObjectName("stop")
        self.gridLayout.addWidget(self.stop, 9, 2, 1, 1)
        self.stop.setEnabled(False)

        ############### logo ADENTIS #####################
        self.logo_adentis = QtWidgets.QLabel(self.central_widget)
        self.logo_adentis.setStyleSheet(
            "background-color: qlineargradient(spread:pad, x1:0.518, y1:0, x2:0.517, y2:1, "
            "stop:0 rgba(186, 230, 255, 0), stop:0.995025 rgba(250, 250, 250, 0));")
        self.logo_adentis.setText("")
        self.logo_adentis.setPixmap(QtGui.QPixmap("Adentis_small2.png"))  # Logo dans le dossier de l'application
        self.logo_adentis.setObjectName("logo_adentis")
        self.gridLayout.addWidget(self.logo_adentis, 2, 2, 2, 1, QtCore.Qt.AlignHCenter)

        ###### Recherche des Valeurs par défauts enregistrées############
        self.params = {'login:': "", 'password:': "", 'xlsx_path:': "", 'audio:': "", 'progress_quit:': "",
                       'option_profil:': "", 'recherche_type:': ""}
        self.get_params()

        self.retranslateUi()
        self.show()

    def select_recherche_type(self, statut):
        sender = self.central_widget.sender()
        if statut is True:
            self.recherche_type = sender.objectName()
            if sender.objectName() == 'Actuel':
                self.recherche_type_both_Act.setChecked(False)
                self.recherche_type_auparavant_Act.setChecked(False)
            elif sender.objectName() == 'Auparavant':
                self.recherche_type_actuel_Act.setChecked(False)
                self.recherche_type_both_Act.setChecked(False)
            elif sender.objectName() == 'Actuel+Auparavant':
                self.recherche_type_actuel_Act.setChecked(False)
                self.recherche_type_auparavant_Act.setChecked(False)
        else:
            sender.setChecked(True)  # On n'autorise pas le déchochage
            self.recherche_type = sender.objectName()  # le type de recherche reste sur l'objet toujours coché
        # print(self.recherche_type)

    def get_params(self):
        if os.path.exists(os.path.join(os.getcwd(), 'config.txt')):
            with open(os.path.join(os.getcwd(), 'config.txt'), 'r') as f:
                for line in f:
                    line = line.strip()
                    self.params.update({param: line[line.index(param) + len(param):] for param in self.params.keys()
                                        if param in line})

        self.identifiant.setText(self.params['login:'])
        self.password.setText(self.params['password:'])
        self.output_file.setText(self.params['xlsx_path:'])
        if self.params['audio:'] == 'True':
            self.audioAct.setChecked(True)
        else:
            self.audioAct.setChecked(False)git
        if self.params['progress_quit:'] == 'True':
            self.progress_auto_quit_Act.setChecked(True)
        else:
            self.progress_auto_quit_Act.setChecked(False)
        if self.params['option_profil:'] == 'True':
            self.optionProfil.setChecked(True)
        else:
            self.optionProfil.setChecked(False)

        if self.params['recherche_type:'] == 'Actuel':
            self.recherche_type_actuel_Act.trigger()
        elif self.params['recherche_type:'] == 'Auparavant':
            self.recherche_type_auparavant_Act.trigger()
        elif self.params['recherche_type:'] == 'Actuel+Auparavant':
            self.recherche_type_both_Act.trigger()

    def nav_activated(self, text: str):
        """
        Slot permettant de récupérer le str renvoyé par le signal de la combo_box "Choix du navigateur"
        et de le sauver dans navigateur
        :param text:
        """
        self.navigateur = text

    def retranslateUi(self):
        self.setWindowTitle("Recherche Linkedin")
        self.titre_navigateur.setText("Navigateur")

        self.lancement.setText("Lancement")
        self.titre_id_linkedin.setText("Identification Linkedin")
        __sortingEnabled = self.listWidget.isSortingEnabled()
        self.listWidget.setSortingEnabled(True)
        self.listWidget.setSortingEnabled(__sortingEnabled)
        self.listWidget2.setSortingEnabled(True)
        self.listWidget2.setSortingEnabled(__sortingEnabled)
        self.society.setText("Société")
        self.poste_keyword.setText("Poste recherché")
        self.poste_antikeyword.setText("Domaine à éviter")
        self.localisation.setText("Localisation")

        self.stop.setText("STOP")
        self.combo_box.setToolTip("<html><head/><body><p align=\"justify\"><span style=\" "
                                  "font-weight:400;\">Choix du navigateur</span></p></body></html>")
        self.combo_box.setItemText(0, "Firefox")
        self.combo_box.setItemText(1, "Chrome")
        self.titre_recherche.setText("Paramètres de la Recherche")

    def get_keywords(self, nom_fichier: str) -> list:
        """
        Récupère les keywords par ligne dans le fichier texte spécifié dans nom_fichier
        :param nom_fichier:
        :rtype: list
        """
        #
        if os.path.exists(os.path.join(os.getcwd(), nom_fichier)):
            with open(os.path.join(os.getcwd(), nom_fichier), 'r') as f:
                keywords = [l.strip() for l in f]
        else:
            keywords = []
            f = open(os.path.join(os.getcwd(), nom_fichier), 'w')
            f.close()
            self.status_bar.showMessage('Fichier source des mots clés de poste introuvable')
        return keywords

    def get_keywords_checked(self, listobject) -> list:
        """
        Récupère les éléments cochés dans le QlistWidget listobject
        :rtype: list
        :param listobject:
        """
        checked_items = []
        for index in range(listobject.count()):
            if listobject.item(index).checkState() == QtCore.Qt.Checked:
                checked_items.append(listobject.item(index).text())
        return checked_items

    def output(self):
        """
        Affiche une fenêtre de dialogue de recherche de fichier et met à jour self.output_file
        """
        # Choix du fichier Excel
        fname = QtWidgets.QFileDialog.getOpenFileName(QtWidgets.QFileDialog(), 'Choisir fichier', os.getcwd(),
                                                      "Excel files (*.xlsx *.xls)")
        if fname[0]:
            self.output_file.setText(fname[0])

    def bot(self):
        """
        Récupère  self.output_file, self.identifiant, self.password, self.societe_champ, self.localisation_champ,
        self.optionprofil, self.navigateur, et les keywords et antikeywords
        Paramètres minimums :
            self.output_file.text()   existe et est un fichier xls
            self.identifiant.text()  est rempli
            self.password.text()    est rempli
            self.societe_champ.text() est rempli
            Au moins un des keywords de self.listWidget est coché
        Appelle la classe Progression
        Appelle la classe TableTpsReel
        Appelle la classe Bot
            Met en place les connection entre les signaux provenant de Bot et les slot de mise à jour de Progression et
            des différent boutons de l'interface
        """
        if os.path.isfile(self.output_file.text()) and 'xls' in os.path.splitext(self.output_file.text())[1] and len(
                self.identifiant.text()) > 0 and len(self.password.text()) > 0 and len(
            self.societe_champ.text()) > 0 and len(self.get_keywords_checked(self.listWidget)) > 0:

            # Sauve l'identifiant, le mot de passe et le chemin du fichier excel dans le fichier config.txt
            self.save_params()

            self.recherche = Bot(self.identifiant.text(), self.password.text(), self.societe_champ.text(),
                                 self.localisation_champ.text(), self.optionProfil.isChecked(),
                                 self.recherche_type, self.output_file.text(),
                                 self.get_keywords_checked(self.listWidget),
                                 self.get_keywords_checked(self.listWidget2),
                                 self.navigateur)
            self.progress = Progression(self.dim)
            try:
                self.table = TableTpsReel(self.output_file.text())
            except ValueError:
                self.progress.update_statut_recherche('Le Fichier Excel n"est pas correctement formaté, Renseigner '
                                                      'au moins le nom des colonnes dans la première ligne')
                self.done()
                return

            # Signaux de mise à jour du label statut_recherche avec les message provenant de Bot et TableTpsReel
            self.recherche.c.update_statut_recherche[str].connect(self.progress.update_statut_recherche)
            self.table.update_statut_recherche[str].connect(self.progress.update_statut_recherche)
            # Signaux de mise à jour du label statut_page avec les message provenant de Bot
            self.recherche.c.update_statut_page[str].connect(self.progress.update_statut_page)
            self.recherche.c.add_page.connect(self.progress.addpage)
            # Signaux de mise à jour des paramètres de la barre de progression en fonction des infos tirées de Bot
            self.recherche.c.get_max[int].connect(self.progress.set_max)
            self.recherche.c.update_pbar.connect(self.progress.addstep)
            # Signaux de mise à jour de TableTpsReel à partir des infos tirées de Bot
            self.recherche.c.update_table[dict].connect(self.table.remplissage_table)
            # Mets à jour les messages et les états des boutons de l'interface quand Bot est terminé
            self.recherche.finished.connect(self.done)
            # Signaux d'arret de la recherche
            self.stop.clicked.connect(self.stop_recherche)
            self.progress.btn_pause.clicked[bool].connect(self.setpause)
            self.status_bar.showMessage('Recherche en cours')
            # Lance le thread de Bot
            self.recherche.start()
            # Une fois Bot lancé, rend le bouton stop accessible et le bouton lancement inaccessible
            self.stop.setEnabled(True)
            self.lancement.setEnabled(False)
        else:
            self.status_bar.showMessage('Certaines informations requises sont manquantes')

    def save_params(self):
        # Mets à jour le dictionnaire avec les données rentrées par l'utilisateur
        self.params['login:'] = self.identifiant.text()
        self.params['password:'] = self.password.text()
        self.params['xlsx_path:'] = self.output_file.text()

        if self.audioAct.isChecked():
            self.params['audio:'] = 'True'
        else:
            self.params['audio:'] = 'False'

        if self.progress_auto_quit_Act.isChecked():
            self.params['progress_quit:'] = 'True'
        else:
            self.params['progress_quit:'] = 'False'

        if self.optionProfil.isChecked():
            self.params['option_profil:'] = 'True'
        else:
            self.params['option_profil:'] = 'False'

        self.params['recherche_type:'] = self.recherche_type

        if os.path.exists(os.path.join(os.getcwd(), 'config.txt')):
            with open(os.path.join(os.getcwd(), 'config.txt'), 'w') as fw:
                for param in self.params.keys():
                    fw.write('{}{}\n'.format(param, self.params[param]))

    def stop_recherche(self):
        self.recherche.continueflag = False
        self.stop.setEnabled(False)

    def setpause(self, pressed):
        if pressed:
            self.progress.btn_pause.setIcon(QtGui.QIcon('playbuttonb.png'))
            self.progress.btn_pause.setIconSize(QtCore.QSize(45, 45))
            self.recherche.pauseflag = True
        else:
            self.progress.btn_pause.setIcon(QtGui.QIcon('pausebutton.png'))
            self.progress.btn_pause.setIconSize(QtCore.QSize(45, 45))
            self.recherche.pauseflag = False

    def done(self):
        """
        Show the message that fetching posts is done.
        Disable Stop button, enable the Start one and reset progress bar to 0
        """
        self.stop.setEnabled(False)
        self.lancement.setEnabled(True)
        self.progress.pbar.setValue(0)
        self.status_bar.showMessage('Recherche terminée')
        if self.progress_auto_quit_Act.isChecked():
            self.progress.close()
        if self.audioAct.isChecked():
            sound = glob(os.path.join(os.getcwd(), '*.mp3'))
            if len(sound) > 0 and os.path.isfile(sound[0]):
                mixer.init()
                mixer.music.load(sound[0])
                mixer.music.play()
            else:
                winsound.Beep()

    def closeEvent(self, e):
        if hasattr(self, 'recherche'):
            self.stop.clicked.connect(self.stop_recherche)
            while self.recherche.isFinished() is False:
                time.sleep(1)
        if hasattr(self, 'table'):
            self.table.close()
        if hasattr(self, 'progress'):
            self.progress.close()
        e.accept()


class QListWidget(QtWidgets.QListWidget):
    def contextMenuEvent(self, e):
        cmenu = QtWidgets.QMenu(self)
        cmenu.setStyleSheet("background-color: rgb(255, 255, 255);\n"
                            "selection-background-color : rgb(0,0,255);")
        newAct = cmenu.addAction('Nouveau Mot clé')
        delAct = cmenu.addAction('Supprimer mot clé')
        action = cmenu.exec_(self.mapToGlobal(e.pos()))

        if self.objectName() == 'listWidget':
            nomfichier = 'keywords-poste.txt'
        elif self.objectName() == 'listWidget2':
            nomfichier = 'antikeywords-poste.txt'

        # Définit un context menu pour les list widget
        if action == newAct:
            item = QtWidgets.QListWidgetItem()
            item.setFlags(
                QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsUserCheckable | QtCore.Qt.ItemIsEnabled)
            item.setCheckState(QtCore.Qt.Checked)
            keyword, ok = QtWidgets.QInputDialog.getText(self, 'Nouveau mot clé', 'Mot clé')
            if ok:
                item.setText(keyword)
                self.addItem(item)
                if os.path.exists(os.path.join(os.getcwd(), nomfichier)):
                    with open(os.path.join(os.getcwd(), nomfichier), 'a') as f:
                        f.write('\n{}'.format(keyword))

        elif action == delAct:
            listItems = self.selectedItems()
            if listItems:
                for item in listItems:
                    self.takeItem(self.row(item))
                    keyword = item.text()
                    if len(keyword) > 0:
                        noms_a_effacer = keyword.encode('utf-8')
                        with open(nomfichier, 'rb') as f:
                            with open(nomfichier, 'r+b') as g:
                                ch = f.read()
                                x = ch.find(noms_a_effacer)
                                x = ch[0:x].rfind(b'\n') + 1
                                f.seek(x)
                                g.seek(x)
                                [g.write(ln) for ln in f if noms_a_effacer not in ln]
                                g.truncate()


############### Classe Interface Barre de progression #####################
class Progression(QtWidgets.QWidget):
    """
    Definit une fenetre avec deux lignes :
        un label statut_recherche
        un label statut_page
    et une barre de progression
    """

    def __init__(self, dim):
        super().__init__()
        width = 1300
        height = 160
        self.setGeometry(455,
                         dim.height() - (height + 50), width, height)
        self.setWindowTitle('Recherche...')
        icon = QtGui.QIcon()
        # Icon dans le dossier de l'application
        icon.addPixmap(QtGui.QPixmap("Adentis_icon.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.setWindowIcon(icon)

        grid = QtWidgets.QGridLayout()
        grid.setSpacing(10)
        self.setLayout(grid)

        # Statut de la recherche
        self.statut_recherche = QtWidgets.QLabel('...', self)
        grid.addWidget(self.statut_recherche, 0, 0)

        # Bouton pause
        self.btn_pause = QtWidgets.QPushButton('', self)
        self.btn_pause.setIcon(QtGui.QIcon('pausebutton.png'))
        self.btn_pause.setIconSize(QtCore.QSize(45, 45))
        self.btn_pause.resize(self.btn_pause.sizeHint())
        self.btn_pause.setCheckable(True)
        # self.btn_pause.clicked[bool].connect(self.setpause)
        grid.addWidget(self.btn_pause, 2, 2)

        # Numéro de la page
        self.page_count = 1
        self.statut_page = QtWidgets.QLabel('Page ' + str(self.page_count), self)
        grid.addWidget(self.statut_page, 1, 0)

        # Barre de progression
        self.pbar = QtWidgets.QProgressBar(self)
        self.pbar.setValue(0)
        self.pbar.setMaximum(10)
        self.pbar.setObjectName("progressBar")
        grid.addWidget(self.pbar, 2, 0, 2, 1)

        self.show()

    def set_max(self, val: int):
        """
        Met à jour la valeur max de la barre de progression
        :param val:
        """
        self.pbar.setMaximum(val)

    def addstep(self):
        """
        Incrémente de 1 la barre de progression
        """
        if self.pbar.value() < self.pbar.maximum():
            self.pbar.setValue(self.pbar.value() + 1)
        else:
            self.pbar.setValue(0)

    def addpage(self):
        """
        Incrémente de 1 le statut de la page
        """
        self.page_count += 1
        self.statut_page.setText('Page ' + str(self.page_count))

    def update_statut_recherche(self, text: str):
        """
        Met à jour le text du label statut_recherche
        :param text:
        """
        self.statut_recherche.setText(text)

    def update_statut_page(self, text: str):
        """
        Met à jour le text du label statut_page
        :param text:
        """
        self.statut_page.setText(text)


##################### Classe Interface Tableau Tps réel ############################
class TableTpsReel(QtWidgets.QWidget):
    """
    Definit une fenetre avec un tableau qui reproduit ce qui est ajouté au fichier
    excel
    """
    update_statut_recherche = pyqtSignal(str)

    def __init__(self, chemin: str):
        super().__init__()
        width = 1300
        height = 790
        self.setGeometry(455, 38, width, height)
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap("Adentis_icon.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.setWindowIcon(icon)
        self.setWindowTitle("Résultat")

        self.gridLayout = QtWidgets.QGridLayout(self)
        self.gridLayout.setContentsMargins(0, 0, 0, 0)
        self.gridLayout.setObjectName("gridLayout")

        self.tableWidget = QtWidgets.QTableWidget(self)
        self.tableWidget.setObjectName("tableWidget")
        sh = load_workbook(chemin)  # Ouvre le fichier
        sh = sh[sh.sheetnames[0]]  # Accède à la première feuille
        self.tableWidget.setColumnCount(sh.max_column)
        self.tableWidget.setRowCount(0)

        self.remplissage_table_init(sh)
        self.gridLayout.addWidget(self.tableWidget, 0, 0, 1, 1)
        self.show()

    def remplissage_table_init(self, sh):
        """
        Recopie les valeurs du fichier excel dans le tablewidget
        :param sh:
        """
        # Teste s'il y a des noms de colonnesd première ligne
        if len(sh.cell(row=1, column=1).value) > 0 or len(sh.cell(row=1, column=2).value) > 0:
            for colnum in range(1, sh.max_column + 1):
                # Inscrit les valeurs de la première ligne comme nom de colonne dans le table widget
                item = QtWidgets.QTableWidgetItem(sh.cell(row=1, column=colnum).value)
                self.tableWidget.setHorizontalHeaderItem(colnum - 1, item)
            # Si le fichier excel a déjà des données (à partir de la deuxième ligne)
            if sh.max_row > 1:
                for rownum in range(2, sh.max_row + 1):
                    self.tableWidget.insertRow(self.tableWidget.rowCount())  # Ajoute une ligne au tableWidget
                    for colnum in range(1, sh.max_column):
                        # Remplit le table Widget à partir du fichier excel
                        item = QtWidgets.QTableWidgetItem(sh.cell(row=rownum, column=colnum).value)
                        self.tableWidget.setItem(self.tableWidget.rowCount() - 1, colnum - 1, item)
            # Fait défiler le tableau jusqu'à la dernière ligne
            self.tableWidget.scrollToBottom()

            # colidx apparie les mot clé des noms de colonne et son numéro dans le tableau.
            self.colidx = dict()
            # colname détermine les mots clés des nom de colonne qui seront recherchés dans le fichier excel
            self.colname = ['société', 'domain', 'nom', 'profil', 'fonction', 'tel', 'mail', 'localisation', 'site']
            # loop through headers and find column number for given column name
            headercount = self.tableWidget.columnCount()
            for x in range(0, headercount, 1):
                headertext = self.tableWidget.horizontalHeaderItem(x).text()
                self.colidx.update({columnname: x for columnname in self.colname if
                                    self.supprime_accent(columnname).upper() in self.supprime_accent(
                                        headertext).upper()})
        else:
            raise ValueError  # Si pas de noms de colonne dans la première ligne retourne une erreur

    def remplissage_table(self, info: dict):
        """
        Remplit le table Widget avec les informations contenues dans le dictionnaires retourné par Bot
        :param info:
        """
        try:
            self.tableWidget.insertRow(self.tableWidget.rowCount())  # insert une nouvelle ligne à la fin du tableau
            for col in self.colidx.keys():
                if col in info.keys():  # si le mot clé dans le nom de colonne correspond à une des clé du dict info
                    # on ajoute l'info correspondante à la bonne place dans la bonne colonne
                    item = QtWidgets.QTableWidgetItem(info[col])
                    self.tableWidget.setItem(self.tableWidget.rowCount() - 1, self.colidx[col], item)
            self.tableWidget.scrollToBottom()  # pour que la dernière ligne reste toujours visible
        except:
            self.update_statut_recherche.emit('Problème dans l"écriture des infos dans le tableau')

    def supprime_accent(self, ligne: str):
        """ supprime les accents du texte source """
        accents = {'a': ['à', 'ã', 'á', 'â'],
                   'e': ['é', 'è', 'ê', 'ë'],
                   'i': ['î', 'ï'],
                   'u': ['ù', 'ü', 'û'],
                   'o': ['ô', 'ö'],
                   ' ': ['-', '_']}
        for (char, accented_chars) in accents.items():
            for accented_char in accented_chars:
                ligne = ligne.replace(accented_char, char)
        return ligne


##############################Bot class ###########################################
####################################################################################

class Communicate(QObject):
    """
    Classe de signaux permettant de mettre à jour les barre de progression et tableau
    """
    update_statut_recherche = pyqtSignal(str)
    update_statut_page = pyqtSignal(str)
    update_pbar = pyqtSignal()
    update_table = pyqtSignal(dict)
    add_page = pyqtSignal()
    get_max = pyqtSignal(int)


class Bot(QThread):
    """
    Classe Bot qui lance le bot et lance la recherche sur linkedin pour scanner les profil résultant de la recherche
    paramètres : identifiant, mdp, societe, region, into_profil, recherche_contrainte, chemin, keywords, antikeywords, navigateur
    """

    def __init__(self, identifiant, mdp, societe, region, into_profil, recherche_type, chemin, keywords, antikeywords,
                 navigateur):
        super().__init__()
        # Infos rentrées dans l'IHM
        self.inputUser = dict()
        self.inputUser['USER'] = identifiant
        self.inputUser['PASSWORD'] = mdp
        self.inputUser['REGION'] = region
        self.inputUser['ENTREPRISE'] = societe
        self.inputUser['intoProfile'] = into_profil
        self.inputUser['recherche_type'] = recherche_type
        self.keywords = keywords
        self.antikeywords = antikeywords
        self.chemin = chemin
        self.nav = navigateur
        self.c = Communicate()
        self.continueflag = True
        self.pauseflag = False
        self.targetedInfo = ['nom', 'fonction', 'société', 'profil', 'localisation', 'tel', 'mail', 'site', 'domain',
                             'auparavant']
        # Ouvre un navigateur en tâche de fond selon le choix de l'utilisateur
        # Ajoute le driver du navigateur dans la variable d'environnement PATH si elle n'y est pas
        if self.nav == 'Firefox':
            if os.path.join(os.getcwd(), 'geckodriver') not in os.environ['PATH']:
                os.environ['PATH'] = os.environ['PATH'] + ';' + os.path.join(os.getcwd(), 'geckodriver')
            options = webdriver.FirefoxOptions()
            options.set_headless(headless=True)
            self.browser = webdriver.Firefox(options=options)
            self.browser.implicitly_wait(1)

        elif self.nav == 'Chrome':
            if os.path.join(os.getcwd(), 'chromedriver') not in os.environ['PATH']:
                os.environ['PATH'] = os.environ['PATH'] + ';' + os.path.join(os.getcwd(), 'chromedriver')
            options = webdriver.ChromeOptions()
            options.set_headless(headless=True)
            self.browser = webdriver.Chrome(options=options)
            self.browser.implicitly_wait(1)

    def run(self):
        while True:
            # Méthode lancé quand on commence le thread
            login = self.login_linkedin()  # Lance une tentative de login sur la page de linkedin
            if login is True:
                self.c.update_statut_recherche.emit("Identification réussie")

                # Point de contrôle pour stopper la recherche si l'utilisateur a appuyé sur STOP
                if self.continueflag is False:
                    self.c.update_statut_recherche.emit("Recherche stoppée par l'utilisateur")
                    break

                # Lance une recherche sur l'entreprise pour en extraire le domaine
                try:
                    self.browser.get(self.create_url_to_search(self.inputUser['ENTREPRISE'], typeSearch='companies'))
                except:
                    self.c.update_statut_recherche.emit('Erreur lors du chargement de la page')
                    break

                if self.continueflag is False:
                    self.c.update_statut_recherche.emit("Recherche stoppée par l'utilisateur")
                    break

                self.waiting_for('search/results/')  # Attend que la page se charge et que l'url est changé
                self.domain, nom_entreprise = self.get_company_domain()  # Extrait le domaine de l'entreprise

                if nom_entreprise == 'OK':
                    # Lance une recherche sur les employés de cette entreprise
                    try:
                        self.browser.get(
                            self.create_url_to_search(self.inputUser['ENTREPRISE'], self.inputUser['REGION'],
                                                      typeSearch='people'))
                        # self.waiting_for('search/results/')
                        wait = self.wait_for_full_loading()
                        # print(self.browser.current_url)

                        self.filtre_entreprise_actuelle()

                        # Lance le bot
                        self.view_bot()  # Lance la récupération des infos de la recherche
                        break  # puis sort de la boucle pour fermer correctement le thread

                    except Exception as exception:
                        # Si une erreur est détectée,
                        # renvoie la ligne et le type d'erreur dans le code puis sort proprement de l'appli
                        # Facilite le débogage sans faire planter l'appli et sans laisser le thread ouvert
                        self.c.update_statut_recherche.emit(
                            "line {} {} {}".format(exception.__traceback__.tb_lineno,
                                                   type(exception).__name__,
                                                   exception.with_traceback(exception.__traceback__)))
                        self.sleep(5)  # Laisse le temps à l'utilisateur de lire le message d'erreur
                        break
                else:
                    self.c.update_statut_recherche.emit('Aucune correspondance avec l"entreprise recherchée')
                    break
            else:
                break
        self.browser.quit()  # Ferme le navigateur

    def login_linkedin(self):
        """
        Login sur linkedin
        :return:
        """
        try:
            self.browser.get('https://www.linkedin.com/uas/login')
        except WebDriverException:
            self.c.update_statut_recherche.emit(
                'Impossible d"atteindre la page d"identification, Vérifier la connexion')
            return False

        self.waiting_for2('https://www.linkedin.com/uas/login')
        if self.browser.current_url == 'https://www.linkedin.com/uas/login':
            emailElement = self.browser.find_element_by_id("session_key-login")
            emailElement.send_keys(self.inputUser['USER'])
            passElement = self.browser.find_element_by_id('session_password-login')
            passElement.send_keys(self.inputUser['PASSWORD'])
            passElement.submit()
            # Waiting for the home page to load with timeout
            self.c.update_statut_recherche.emit('Identification...')
            self.waiting_for2("https://www.linkedin.com/feed/?trk=")
            if self.browser.current_url == "https://www.linkedin.com/feed/?trk=":
                return True
            else:
                self.c.update_statut_recherche.emit(
                    'L"identification a échouée, vérifier les identifiants et mots de passe')
                return False
        else:
            self.c.update_statut_recherche.emit(
                'La durée de connexion à la page est anormalement longue, vérifiez votre connexion')
            return False

    def create_url_to_search(self, *keywords, typeSearch: str):
        """
        Create the url of the research based on the user input
        :param keywords:
        :param typeSearch:
        :rtype str:
        """
        basic_url_start = 'https://www.linkedin.com/search/results/' + typeSearch + '/?keywords='
        basic_url_end = '&origin=GLOBAL_SEARCH_HEADER'
        # if typeSearch == "people":
        #     basic_url_end = basic_url_end + '&page=34'
        for keyword in keywords:
            basic_url_start = basic_url_start + '%20' + keyword.replace(' ', '%20')
        self.c.update_statut_recherche.emit("[+] Searching for : {}".format(keywords))
        return basic_url_start + basic_url_end

    def get_company_domain(self):
        """
        Extrait les informations de domaine de l'entreprise recherchée
        :return:
        """
        page = BeautifulSoup(self.browser.page_source, "html.parser")
        nom_entreprise = 'Aucune correspondance'
        domain = ""
        for link in page.find_all('div', class_='search-result__info pt3 pb4 pr0'):
            try:
                if self.supprime_accent(self.inputUser['ENTREPRISE']).upper() in self.supprime_accent(
                        link.select('a > h3')[0].text.strip()).upper():
                    nom_entreprise = 'OK'
                    domain = link.select('p.subline-level-1.Sans-15px-black-85%.search-result__truncate')[
                        0].text.strip()
                    # print(domain)
                    if domain is not False:
                        break
            except:
                break
        return domain, nom_entreprise

    def filtre_entreprise_actuelle(self):
        """
        Clique sur le bouton Tous les filtres de la page de recherche et sélectionne les champs utiles selon
        le choix d'option de l'utilisateur 'Actuel, Auparavant, Actuel+Auparavant'
        et Active le filtre sur la localisation quand c'est possible
        :return:
        """
        try:
            self.browser.execute_script("window.scrollTo(0, 0);")  # s'assure qu'on est bien en haut de la page
            # Clique sur le bouton des filtres avancés
            self.browser.find_element_by_xpath(
                "//button[@class='search-filters-bar__all-filters button-tertiary-medium-muted mr3']").click()
            # Récupère le conteneur où sont les optiops de filtres
            form = self.browser.find_element_by_xpath("//div[@class='search-advanced-facets__layout display-flex ph0']")
            # Récupère le conteneur de l'entête où sont les boutons pour appliquer les filtres
            form_header = self.browser.find_element_by_xpath(
                "//div[@class='search-advanced-facets__layout display-flex align-items-center justify-space-between']")

            # Si l'utilsateur a choisi le mode Actuel+Auparavant, on ne fait rien c'est la recherche normale

            if self.inputUser['recherche_type'] == 'Actuel':  # Si l'utilisateur à choisi le mode Actuel
                # Localise le conteneur des options du filtre Entreprise Actuelle
                zoneactuel = form.find_element_by_xpath(
                    ".//fieldset[@class='search-s-facet__values search-s-facet__values--facetCurrentCompany']")
                self.browser.execute_script("arguments[0].scrollIntoView(true);", zoneactuel)
                # récupère la division blocante pour sélectionner les entreprises
                block = zoneactuel.find_element_by_xpath(
                    ".//ol[@class='search-s-facet__list list-style-none']")

                # Récupère les propositions de Linkedin
                listentreprise = zoneactuel.find_elements_by_xpath(".//li[@class='search-facet__value ']")
                for entreprise in listentreprise:
                    entrepriseid = entreprise.find_element_by_xpath(
                        ".//label[@class='search-s-facet-value__label Sans-15px-black-70%']")
                    # print(entrepriseid.text)

                    if self.supprime_accent(self.inputUser['ENTREPRISE']).upper() in self.supprime_accent(
                            entrepriseid.text).upper():
                        # Récupère l'élément cliquable
                        entreprise.find_element_by_xpath(".//input[@class='medium-input mr3']")
                        # Clique sur la checkbox de l'entreprise en passant au delà de la division blocante
                        webdriver.ActionChains(self.browser).move_to_element(block).click(entreprise).perform()
                        break

            if self.inputUser['recherche_type'] == 'Auparavant':  # Si l'utilisateur a choisi le mode auparavant
                zoneauparavant = form.find_element_by_xpath(
                    ".//fieldset[@class='search-s-facet__values search-s-facet__values--facetPastCompany']")
                self.browser.execute_script("arguments[0].scrollIntoView(true);", zoneauparavant)
                # self.sleep(2)
                block = zoneauparavant.find_element_by_xpath(
                    ".//ol[@class='search-s-facet__list list-style-none']")  # récupère la division blocante pour sélectionner les entreprises
                listentreprise = zoneauparavant.find_elements_by_xpath(".//li[@class='search-facet__value ']")
                for entreprise in listentreprise:
                    entrepriseid = entreprise.find_element_by_xpath(
                        ".//label[@class='search-s-facet-value__label Sans-15px-black-70%']")
                    # print(entrepriseid.text)
                    if self.supprime_accent(self.inputUser['ENTREPRISE']).upper() in self.supprime_accent(
                            entrepriseid.text).upper():
                        # Récupère l'élément cliquable
                        entreprise.find_element_by_xpath(".//input[@class='medium-input mr3']")
                        # Clique sur la checkbox de l'entreprise en passant au delà de la division blocante
                        webdriver.ActionChains(self.browser).move_to_element(block).click(entreprise).perform()
                        self.sleep(1)
                        break

            try:
                # Récupère dans l'input de localisation et récherche :région de ville et en extrait la ville
                # préserve la possibilité d'utiliser d'autres termes de recherches dans la partie localisation
                # de l'interface
                region = re.search("REGION DE " + r"[\w]+", self.supprime_accent(self.inputUser['REGION']).upper())
                if region:
                    ville = region.group().replace('REGION DE ', "")

                    # Applique le filtre sur la zone géographique si proposé par linkedin
                    zonelieu = form.find_element_by_xpath(
                        ".//fieldset[@class='search-s-facet__values search-s-facet__values--facetGeoRegion']")
                    self.browser.execute_script("arguments[0].scrollIntoView(true);", zonelieu)
                    # self.sleep(2)

                    # récupère la division blocante pour sélectionner les entreprises
                    block = zonelieu.find_element_by_xpath(
                        ".//ol[@class='search-s-facet__list list-style-none']")
                    listlieu = zonelieu.find_elements_by_xpath(".//li[@class='search-facet__value ']")

                    for lieu in listlieu:
                        lieuid = lieu.find_element_by_xpath(
                            ".//label[@class='search-s-facet-value__label Sans-15px-black-70%']")
                        # print(entrepriseid.text)

                        if ville in self.supprime_accent(
                                lieuid.text).upper():
                            # Récupère l'élément cliquable
                            lieu.find_element_by_xpath(".//input[@class='medium-input mr3']")
                            # Clique sur la checkbox de l'entreprise en passant au delà de la division blocante
                            webdriver.ActionChains(self.browser).move_to_element(block).click(lieu).perform()
                            break

            except exceptions.NoSuchElementException:  # Si la partie filtre n'est pas accessible
                self.c.update_statut_recherche.emit(
                    "Le filtre région n'est pas disponible")
                self.sleep(5)  # laisse le temps à l'utilisateur de lire le message d'erreur

            # clique sur le bouton pour appliquer les filtres
            form_header.find_element_by_xpath(
                ".//button[@class='search-advanced-facets__button--apply button-primary-large']").click()
            self.sleep(5)  # laisse le temps à l'utilisateur de lire le message d'erreur
            self.wait_for_full_loading()
        except:
            self.c.update_statut_recherche.emit(
                'Problème dans la sélection des filtres, Actuel + Auparavant par défaut')
            self.inputUser['recherche_type'] = 'Actuel+Auparavant'
            self.sleep(5)  # laisse le temps à l'utilisateur de lire le message d'erreur

    def view_bot(self):
        """
        Coeur du Bot qui appelle les autres fonctions
        :return:
        """
        visited = self.get_visited()
        count = 0  # Comptabilise le nombre de profil acceptés
        pidx = 1  # Compteur de page
        page = BeautifulSoup(self.browser.page_source, "html.parser")

        while self.continueflag:  # Boucle jusqu'à la dernière page de la

            # sauve l'url de la page de recherche pour y retourner si on rentre dans les profils
            search_page = self.browser.current_url

            self.pause()  # vérifie le pauseflag et met en pause si == True

            people, job, loc, statut = self.get_people_links(page, visited)
            # récupère les liens des personnes correspondants aux keywords et n'étant pas déjà dans le fichier excel

            if people:
                self.c.get_max.emit(len(people))
                for nom, person in people.items():  # Pour chaque profil correspondant au critères
                    # print(nom)
                    # print(statut[nom])
                    # print(job[nom])
                    self.pause()  # vérifie le pauseflag et met en pause si == True

                    if self.continueflag is False:  # Si bouton STOP préssé pendant la boucle, sors de la boucle
                        break

                    info = dict()  # initialise le dictionnaire qui va receuillir les infos du profil en cours

                    time.sleep(random.uniform(3.5, 6.9))  # add random to make us look human

                    ID = self.get_id(people[nom])  # construit l'URL complète du profil

                    count += 1

                    # Récupère les infos présentes dans la page de recherche
                    info = self.get_list_info(nom, job, ID, loc, statut[nom])
                    # Mets à jour le statut de recherche dans la barre de progression
                    self.c.update_statut_recherche.emit(
                        "[+]" + nom + " checked \n (" + str(count) + ") Checked")

                    if self.inputUser['intoProfile'] is True or info['statut'] == 'Indéterminé':

                        self.browser.get(ID)  # va sur le profil

                        self.waiting_for2("/in/")  # Attend le chargement de la page

                        if "/in/" in self.browser.current_url:
                            info = self.get_profile_infos(info)  # Récupère les infos sur le profil des person

                            # Mets à jour le statut de recherche dans la barre de progression
                            self.c.update_statut_recherche.emit("Profil visited!")
                        else:
                            # si le lien du profil ne renvoie à rien (page "Utilisateur Linkedin non accessible")
                            # on passe au lien suivant
                            continue

                    self.c.update_pbar.emit()  # met à jour la barre de progression
                    # N'inscrit dans la tableau que si c'est un poste actuel
                    if info['statut'] == 'Actuel':
                        # Remplit le champ domaine uniquement quand c'est l'entreprise actuelle
                        info['domain'] = self.domain
                        self.c.update_table.emit(info)  # Remplit le tableau

                    self.transfer_excel(info)  # Transfert les données récoltée sur excel
                    visited.add(nom)  # Ajoute le nouveau nom à la liste des profils visités

            else:  # S'il n'ya personne de retenu dans la page de recherche
                self.c.update_statut_recherche.emit('Aucune correspondance')

            if self.browser.current_url != search_page:
                self.browser.get(search_page)  # retourne à la page des recherche

            wait = self.wait_for_full_loading()  # Attend le chargement de la page
            if wait == 1:
                raise exceptions.TimeoutException('Temps de chargement de la page anormalement long')

            self.pause()  # vérifie le pauseflag et met en pause si == True

            try:
                self.browser.find_element_by_xpath("//button[@class='next']").click()  # Clique sur le bouton next
                pidx += 1
                self.waiting_for("&page=" + str(pidx))  # Attend le chargement de la page
                wait = self.wait_for_full_loading()  # Attend le chargement de la page
                if wait == 1:
                    raise exceptions.TimeoutException('Temps de chargement de la page anormalement long')

                # print(self.browser.current_url)
                if self.browser.current_url == search_page:  # Si l'url n'a pas changé après next, sort de la boucle
                    self.c.update_statut_page.emit('Passement de page défaillant')
                    break  # Pour que le message soit le dernier vu par l'utilisateur et non pas les message de except

                page = BeautifulSoup(self.browser.page_source, "html.parser")  # analyse le code de la nouvelle page
                self.c.add_page.emit()  # Mets à jour le statut page de la barre de progression

            except:  # Si on échoue à toruver le bouton next
                self.c.update_statut_page.emit('Dernière page atteinte')
                self.c.update_statut_recherche.emit('Plus personne à voir')
                self.sleep(5)
                break

        if self.continueflag is False:  # Si la recherche a été stoppé par l'utilisateur
            self.c.update_statut_recherche.emit("Recherche stoppée par l'utilisateur")

        return

    def get_visited(self):
        """
        Lit le fichier Excel, récupère les noms déjà inclus dans la liste des noms
        :return: visited
        """
        visited = set()
        sh = load_workbook(self.chemin)
        # Cherche dans la première feuille les profils visités
        sh = sh[sh.sheetnames[0]]
        colname = self.get_excel_columnid(sh)
        for rownum in range(1, sh.max_row + 1):
            visited.add(str(sh.cell(row=rownum, column=colname['nom']).value))
        ## Cherche dans la 2ème feuille les
        # sh = sh[sh.sheetnames[1]]
        # colname = self.get_excel_columnid(sh)
        # for rownum in range(1, sh.max_row + 1):
        #     visited.add(str(sh.cell(row=rownum, column=colname['nom']).value))
        return visited

    def get_id(self, url: str):
        """
        Constuit l'url complète du profil linkedin
        :param url:
        :return:
        """
        root = 'http://www.linkedin.com'
        return root + url

    def get_people_links(self, page, visited):
        """
        Récupère les liens dans la page de recherche selon la concordance du poste actuel avec les keywords et les antikeywords
        :rtype dict:
        """
        links = dict()
        job = dict()
        loc = dict()
        statut = dict()

        for link in page.find_all('div', class_='search-result__info pt3 pb4 ph0'):
            lien = link.find('a')
            url = lien.get('href')
            # print(url)
            poste1 = ""  # poste Actuel si disponible
            poste2 = ""  # poste description
            lieu = ""  # localisation du client
            if url:
                if '/in/' in url and url not in links.values():
                    nom = link.find('span', class_='name actor-name').text
                    # Cherche le poste actuel
                    try:
                        poste1 = link.find('p',
                                           class_='search-result__snippets mt2 Sans-13px-black-55% ember-view').text
                        if 'Actuel :' in poste1:
                            poste1 = poste1.replace('Actuel :', '').strip()
                        else:
                            poste1 = ""

                    except:
                        pass
                    # Si le poste actuel n'est pas disponible, cherche dans le titre du client
                    try:
                        poste2 = link.select('p.subline-level-1.Sans-15px-black-85%.search-result__truncate')[
                            0].text.strip()
                    except:
                        pass
                    # Cherche le lieu de travail, Pour vérifier que cela correspond bien à la localisation recherchée
                    try:
                        lieu = link.select('p.subline-level-2.Sans-13px-black-55%.search-result__truncate')[
                            0].text.strip()
                    except:
                        pass

                    poste = poste1 + ' ' + poste2  # On concatène les deux lignes pour être sur d'avoir le maximum de motsclés
                    alerte_anti_key = 0  # initialise la variable
                    for antikeyword in self.antikeywords:
                        if poste:
                            if self.supprime_accent(
                                    antikeyword).upper() in self.supprime_accent(poste).upper():
                                # Dès qu'au moins un des anti keyword est détecté on flag et on sort de la boucle
                                alerte_anti_key = 1
                                break
                    # print(nom)
                    # print(poste)
                    for keyword in self.keywords:
                        if poste:
                            if self.supprime_accent(keyword).upper() in self.supprime_accent(poste).upper() \
                                    and alerte_anti_key == 0 and nom not in visited:
                                # filtre sur les postes et les noms déjà présents dans le fichier excel
                                if keyword == 'CTO' and 'DIRECTOR' in self.supprime_accent(poste).upper():
                                    continue  # lève une ambiguité

                                links[nom] = url
                                if 'CHEZ ' in self.supprime_accent(
                                        poste1).upper() or ' AT ' in self.supprime_accent(poste1).upper():
                                    job[nom] = poste1  # ligne Poste "Actuel"
                                else:
                                    job[nom] = poste2  # ligne Titre

                                loc[nom] = lieu

                                if self.supprime_accent(
                                        self.inputUser['ENTREPRISE']).upper() in self.supprime_accent(poste).upper():
                                    # Si le nom de l'entreprise est dans les intitulé de poste > Poste actuel
                                    statut[nom] = 'Actuel'
                                else:
                                    if 'CHEZ ' in self.supprime_accent(
                                            poste).upper() or ' AT ' in self.supprime_accent(poste).upper():
                                        # Si on trouve Chez ou At -> il y a mention d'une entreprise -> Nouveau poste
                                        statut[nom] = 'Nouveau'
                                    else:  # Si pas de mention de l'entreprise dans les intitulé de poste
                                        statut[nom] = 'Indéterminé'
                                print(nom)
                                print(job[nom])

                                break

        self.c.update_statut_recherche.emit('Récupération de ' + str(len(links)) + ' liens OK')
        return links, job, loc, statut

    def get_list_info(self, nom: str, job: dict, ID: str, loc: dict, statut: str):
        """
        Extrait les infos directement depuis la page de recherche
        :param nom:
        :param job:
        :param ID:
        :param loc:
        :return:
        """
        info = {target: "" for target in self.targetedInfo}  # initialise les valeurs du dictionnaire
        info['nom'] = nom
        info['fonction'] = job[nom]
        info['localisation'] = loc[nom]
        try:
            boite = job[nom][job[nom].upper().index('CHEZ') + 5:]
        except ValueError:
            try:
                boite = job[nom][job[nom].upper().index(' AT ') + 3:]
            except ValueError:
                boite = self.inputUser['ENTREPRISE'] + ' (défaut)'
        info['société'] = boite
        info['profil'] = ID
        info['statut'] = statut
        return info

    def profile_infos_button(self):
        """
        Essaie de trouver le bouton dépliant permettant d'acceder aux infos personnelles des profils
        :return:
        """
        try:
            # Recherche le bouton sur un profile non connecté à l'utilisateur
            button = self.browser.find_element_by_css_selector('.pv-top-card-v2-section__contact-info')
            button.click()
            time.sleep(2)  # Laisse le temps de charger
            return True
        except:
            self.c.update_statut_recherche.emit("Pas trouvé le bouton")
            time.sleep(3)
            return False

    def get_info_links(self, profilePage, info: dict):
        """
        Récupère les infos personnelles dans le profil et les stocke dans un dict
        :param profilePage:
        :param info:
        :return info:
        """

        # contact = profilePage.select("div.pv-profile-section__section-info.section-info")
        try:
            i = profilePage.select("section.pv-contact-info__contact-type.ci-vanity-url > "
                                   "div.pv-contact-info__ci-container > "
                                   "a")[0]
            info['profil'] = i.get('href')
            # print(info['profil'])
        except:
            pass

        try:
            i = profilePage.select("section.pv-contact-info__contact-type.ci-email > "
                                   "div.pv-contact-info__ci-container > "
                                   "a")[0]
            info['mail'] = i.get('href')
        except:
            pass

        try:
            i = profilePage.select('section.pv-contact-info__contact-type.ci-websites > '
                                   'ul.list-style-none > '
                                   'li.pv-contact-info__ci-container > div > a')[0]
            info['site'] = i.get('href')
            # print(info['site'])
        except:
            pass
        return info

    def get_profile_infos(self, info: dict):
        """
        Récupère les infos disponibles sur la page de profil linkedin et les sauvent dans le dict infos
        Si statut est == indéterminé on recherche dans la partie expérience pour voir le dernier poste et vérifie si
        l'entreprise correspond à l'entreprise recherchée,
        ajoute une nouvelle entrée statut au dictionnaire info pour déterminer si Actuel ou Nouveau
        Si l'option Coordonnées est cochée, on récupère les coordonnées disponibles (dépend de la connexion avec l'utilisateur
        :param info:
        :param statut
        :return:
        """
        # Fais défiler toute la page pour que tous les élements du profil se chargent et soient accessible dans le code
        # source de la page
        for x in range(0, 5000, 5):
            self.browser.execute_script("window.scrollTo(0, {});".format(x))

        profilePage = BeautifulSoup(self.browser.page_source, "html.parser")

        try:
            info['localisation'] = \
                profilePage.select("h3.pv-top-card-section__location.Sans-17px-black-55%-dense.mt1.inline-block")[
                    0].text.strip()
        except:
            info['localisation'] = ''

        if info['statut'] == 'Indéterminé':
            xpcount = 0
            for experience in profilePage.find_all('li',
                                                   class_='pv-profile-section__card-item pv-position-entity ember-view'):
                # print(experience)
                xpcount += 1
                entreprise = experience.select('h4.Sans-17px-black-85% > span.pv-entity__secondary-title')[
                    0].text.strip()

                if xpcount == 1:  # Si c'est le dernier poste en date et donc la première entrée de la section
                    info['société'] = entreprise
                    info['fonction'] = experience.select('h3.Sans-17px-black-85%-semibold')[0].text.strip()
                    if self.supprime_accent(self.inputUser['ENTREPRISE']).upper() in self.supprime_accent(
                            entreprise).upper():
                        info['statut'] = 'Actuel'
                        break  # Si c'est le poste actuel, pas besoin de chercher plus loin, on sort de la boucle
                    else:
                        info['statut'] = 'Nouveau'

                else:  # Cherche le poste que le client occupait quand il travaillait dans l'entreprise recherchée
                    if self.supprime_accent(self.inputUser['ENTREPRISE']).upper() in self.supprime_accent(
                            entreprise).upper():
                        info['auparavant'] = experience.select('h3.Sans-17px-black-85%-semibold')[0].text.strip() \
                                             + ' chez ' + entreprise

        if self.inputUser['intoProfile']:
            button = self.profile_infos_button()
            # Déroule le menu où on peut trouver les infos de contacts
            if button:
                info = self.get_info_links(profilePage, info)
                try:
                    info['tel'] = profilePage.select(
                        "section.pv-contact-info__contact-type.ci-phone > ul.list-style-none > "
                        "li.pv-contact-info__ci-container > span.Sans-15px-black-85%")[
                        0].text.strip()
                    # print(info['tel'])
                except:
                    info['tel'] = 'indisponible'
            else:
                info['profil'] = self.browser.current_url

        return info

    def transfer_excel(self, infos: dict):
        """
        Récupère les infos des profils Linkedin et les écrits dans le fichier Excel choisis par l'utilisateur
        :param infos:
        :return:
        """
        try:
            book = load_workbook(self.chemin)
            if infos['statut'] == 'Actuel':
                # Si le client est actuellement dans l'entreprise
                sheet = book[book.sheetnames[0]]  # On écrit sur la première feuille du fichier excel
            elif infos['statut'] == 'Nouveau':
                # Si le client n'est plus dans l'entreprise mais y a travaillé auparavant
                sheet = book[book.sheetnames[1]]  # On écrit sur la deuxième feuille du fichier excel
            else:
                sheet = book[book.sheetnames[0]]  # Pardéfaut on écrit sur la première feuille

            new_row = sheet.max_row + 1
            colidx = self.get_excel_columnid(sheet)
            linkFont = Font(color=colors.BLUE, underline='single', name='Arial')
            for key in self.targetedInfo:
                if key in colidx.keys():
                    if key == 'profil' or key == 'mail':
                        sheet.cell(row=new_row, column=colidx[key]).hyperlink = infos[key]
                        sheet.cell(row=new_row, column=colidx[key]).font = linkFont
                    else:
                        sheet.cell(row=new_row, column=colidx[key]).value = infos[key]

            book.save(self.chemin)
            status = 0
        except:
            self.c.update_statut_recherche.emit("L'écriture dans le fichier excel a rencontré un problème")
            status = 1
        return status

    def get_excel_columnid(self, sh):
        """
        Lis le fichier Excel et récupère les index des des différentes colonnes
        :param sh:
        """
        # keywords = ['nom', 'fonction', 'société', 'tel', 'mail', 'profil', 'domain', 'localisation','site']
        colnumid = dict()
        for keyword in self.targetedInfo:
            colnumid.update({keyword: colnum for colnum in range(1, sh.max_column + 1) if
                             self.supprime_accent(keyword).upper() in self.supprime_accent(
                                 str(sh.cell(row=1, column=colnum).value)).upper()})
        return colnumid


    def pause(self):
        """
        Vérifie le statut de pauseflag et attend s'il est True
        :return:
        """
        while self.pauseflag:  # tant que progress.btn_pause préssé pauseflag = True
            self.c.update_statut_recherche.emit("Recherche mise en pause par l'utilisateur")
            time.sleep(1)  # tant que pauseflag = True on attend
            if self.continueflag is False:  # pendant la pause, si le bouton STOP est préssé continueflag = False
                break  # si continueflag = False on sort de la boucle

    def waiting_for(self, waitingfor):
        """
        Wait until the waitingfor string is contained in the url
        """
        while waitingfor not in self.browser.current_url:
            time.sleep(random.uniform(3.5, 6.9))

    def waiting_for2(self, waitingfor, timeout=5):
        """
        Wait until the waitingfor string is contained in the url
        """
        count = 0
        while waitingfor not in self.browser.current_url and count < timeout:
            count += 1
            # self.c.update_statut_recherche.emit(str(count))
            time.sleep(3)

    def wait_for_full_loading(self):
        """
        Wait for Full loading
        :return:
        """
        delay = 10
        error = 0
        try:
            self.browser.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            myElem = WebDriverWait(self.browser, delay).until(
                EC.presence_of_element_located((By.ID, 'expanded-footer')))
            # print('Page is ready!')
        except TimeoutException:
            self.c.update_statut_recherche.emit(
                'La durée de connexion à la page est anormalement longue, vérifiez votre connexion')
            error = 1
        return error

    def supprime_accent(self, ligne):
        """ supprime les accents du texte source """
        accents = {'a': ['à', 'ã', 'á', 'â', 'à'.upper(), 'ã'.upper(), 'á'.upper(), 'â'.upper()],
                   'e': ['é', 'è', 'ê', 'ë', 'é'.upper(), 'è'.upper(), 'ê'.upper(), 'ë'.upper()],
                   'i': ['î', 'ï', 'î'.upper(), 'ï'.upper()],
                   'u': ['ù', 'ü', 'û', 'ù'.upper(), 'ü'.upper(), 'û'.upper()],
                   'o': ['ô', 'ö', 'ô'.upper(), 'ö'.upper()],
                   ' ': ['-', '_']}
        for (char, accented_chars) in accents.items():
            for accented_char in accented_chars:
                ligne = ligne.replace(accented_char, char)
        return ligne


if __name__ == "__main__":
    import sys

    app = QtWidgets.QApplication(sys.argv)
    dimScreen = app.desktop().screenGeometry()
    ui = UiMainWindow(dimScreen)
    sys.exit(app.exec_())
